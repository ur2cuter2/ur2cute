from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_OUTPUT = Path("estimate_tool/見積書作成ツール_v0.1.xlsx")
DEFAULT_BRANCH_SOURCE = Path("分岐項目.xlsx")

SHEETS = [
    "00_操作画面",
    "01_案件基本",
    "02_分岐入力",
    "03_見積明細",
    "04_見積書",
    "05_権CSV出力",
    "M_料金表マスタ",
    "M_分岐マスタ",
    "M_受任別分岐",
    "M_連絡不能相続人分岐",
    "M_分岐サマリー",
    "M_見積ロジック",
    "M_CSVマッピング",
    "99_ログ",
]

THIN = Side(style="thin", color="000000")
HEADER_FILL = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")


def normalize_row(row: Iterable) -> List[str]:
    values = ["" if v is None else str(v).strip() for v in row]
    while values and values[-1] == "":
        values.pop()
    return values


def read_sheet_rows(path: Path, sheet_name: str) -> list[list[str]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{path} にシート '{sheet_name}' が見つかりません。")
    ws = wb[sheet_name]
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = normalize_row(row)
        if any(vals):
            rows.append(vals)
    return rows


def write_table(ws, headers: list[str], rows: list[list[str]] | None = None) -> None:
    rows = rows or []
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for row in rows:
        ws.append(row[: len(headers)] + [""] * max(0, len(headers) - len(row)))
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c in r:
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            c.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [14, 12, 40, 14, 10, 14, 28, 22]
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + i)].width = widths[i - 1] if i <= len(widths) else 16


def build_estimate_layout(ws) -> None:
    ws["A1"] = "見積書"
    ws["A1"].font = Font(size=20, bold=True)
    ws.merge_cells("A1:D1")
    ws["A3"] = "見積日"
    ws["C3"] = "有効期限"
    ws["A4"] = "御見積額"
    ws["A3"].font = ws["C3"].font = ws["A4"].font = Font(bold=True)

    headers = ["区分", "種別", "報酬額", "登録免許税又は実費"]
    start = 7
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center")
    for r in range(start + 1, start + 13):
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    summary_labels = ["報酬小計", "実費小計", "合計", "消費税", "源泉徴収税額", "御見積額", "備考"]
    base = start + 14
    for idx, label in enumerate(summary_labels):
        ws.cell(row=base + idx, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=base + idx, start_column=1, end_row=base + idx, end_column=2)
        for c in range(1, 5):
            ws.cell(row=base + idx, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28


def main() -> None:
    parser = argparse.ArgumentParser(description="見積書作成ツール用ワークブックを作成します。")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="出力先 .xlsx")
    parser.add_argument("--branch-source", type=Path, default=DEFAULT_BRANCH_SOURCE, help="分岐項目.xlsx のパス")
    args = parser.parse_args()

    if args.output.suffix.lower() == ".xlsm":
        raise SystemExit(".xlsm は出力できません。--output には .xlsx を指定してください。")
    if args.output.suffix.lower() != ".xlsx":
        raise SystemExit("出力拡張子は .xlsx のみ対応です。")

    args.output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    for s in SHEETS:
        wb.create_sheet(s)

    flow_rows = read_sheet_rows(args.branch_source, "相続業務フロー")
    appoint_rows = read_sheet_rows(args.branch_source, "受任別分岐")
    missing_heir_rows = read_sheet_rows(args.branch_source, "連絡不能相続人分岐")
    summary_rows = read_sheet_rows(args.branch_source, "分岐サマリー")

    ws_in = wb["02_分岐入力"]
    in_headers = ["大項目", "枝番", "内容", "選択", "数量", "金額", "備考"]
    input_rows = [[r[0] if len(r) > 0 else "", r[1] if len(r) > 1 else "", r[2] if len(r) > 2 else "", "未選択", "", "", r[3] if len(r) > 3 else ""] for r in flow_rows]
    write_table(ws_in, in_headers, input_rows)
    dv = DataValidation(type="list", formula1='"未選択,該当,非該当,不明"', allow_blank=False)
    ws_in.add_data_validation(dv)
    dv.add(f"D2:D{max(2, ws_in.max_row)}")

    ws_master = wb["M_分岐マスタ"]
    write_table(ws_master, ["大項目", "枝番", "内容", "備考"], [[r[0] if len(r)>0 else "", r[1] if len(r)>1 else "", r[2] if len(r)>2 else "", r[3] if len(r)>3 else ""] for r in flow_rows])

    write_table(wb["M_受任別分岐"], ["受任内容", "必要作業"], appoint_rows)
    write_table(wb["M_連絡不能相続人分岐"], ["番号", "枝番", "段階", "確認事項・処理", "選択", "次の処理", "関連手続", "備考"], missing_heir_rows)
    write_table(wb["M_分岐サマリー"], ["起点", "分岐", "該当する場合", "主な次アクション", "裁判所手続", "備考"], summary_rows)

    fee_headers = ["fee_id", "分類", "項目名", "単価", "単位", "課税区分", "数量ロジック", "有効フラグ", "備考"]
    fee_rows = [
        ["F001", "登記", "相続登記申請", "", "件", "課税", "案件", 1, ""],
        ["F002", "収集", "戸籍等代理取得", "", "式", "課税", "対象人数", 1, ""],
        ["F003", "収集", "固定資産評価証明書・名寄せ取得", "", "通", "課税", "不動産数", 1, ""],
        ["F004", "書面", "相続関係説明図作成", "", "式", "課税", "案件", 1, ""],
        ["F005", "書面", "法定相続情報一覧図作成", "", "式", "課税", "要否", 1, ""],
    ]
    write_table(wb["M_料金表マスタ"], fee_headers, fee_rows)

    logic_rows = [
        ["L001", "相続登記", "該当", "F001", "1", "明細", "相続登記申請"],
        ["L002", "戸籍収集（代理取得）", "該当", "F002", "数量", "明細", "戸籍等代理取得"],
        ["L003", "評価証明取得", "該当", "F003", "数量", "明細", "固定資産評価証明書・名寄せ取得"],
        ["L004", "相続関係説明図作成", "該当", "F004", "1", "明細", ""],
        ["L005", "法定相続情報一覧図必要", "該当", "F005", "1", "明細", ""],
        ["L006", "協議書方式", "該当", "F006", "1", "明細", "遺産分割協議書作成"],
        ["L007", "証明書方式", "該当", "F007", "1", "加算", "遺産分割協議証明書形式加算"],
        ["L008", "代償金支払あり", "該当", "F008", "1", "加算", "協議書追記事項"],
        ["L009", "売却前提代償金支払", "該当", "F008", "1", "加算", "協議書追記事項"],
        ["L010", "数次相続あり", "該当", "F008", "1", "加算", "協議書追記事項"],
        ["L011", "生命保険調査必要", "該当", "F009", "数量", "明細", "生命保険調査"],
        ["L012", "証券会社口座照会必要", "該当", "F010", "数量", "明細", "証券口座調査"],
        ["L013", "預貯金調査必要", "該当", "F011", "数量", "明細", "預貯金調査または残高証明書"],
        ["L014", "連絡不能相続人あり", "該当", "F012", "数量", "明細", "連絡文作成、発送代行、必要に応じて裁判所手続"],
        ["L015", "成年後見人選任申立", "該当", "F013", "1", "明細", ""],
        ["L016", "相続財産清算人選任申立", "該当", "F014", "1", "明細", ""],
        ["L017", "特別代理人選任申立", "該当", "F015", "1", "明細", ""],
        ["L018", "遺産分割調停", "該当", "F016", "1", "明細", ""],
        ["L019", "抵当権抹消", "該当", "F017", "数量", "明細", ""],
        ["L020", "所有権保存", "該当", "F018", "数量", "明細", ""],
    ]
    write_table(wb["M_見積ロジック"], ["logic_id", "枝番", "条件", "fee_id", "数量ソース", "出力区分", "備考"], logic_rows)

    build_estimate_layout(wb["04_見積書"])

    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
